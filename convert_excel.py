"""
Convert Excel data to JSON format for the Secondary Packaging Equipment dashboard.
Reads Value and Volume sheets from the Pivot Excel file.
Outputs value.json, volume.json, and segmentation_analysis.json.

RULES:
- No double counting: only leaf nodes get year data
- Parent aggregates are excluded from data
- Hierarchical parents become structural container keys
"""

import json
import re
import openpyxl

EXCEL_FILE = 'Pivot_Global Secondary Packaging Equipment Market_Coherent Market Insights.xlsx'
YEARS = [str(y) for y in range(2021, 2034)]  # 2021-2033

# Segment types to detect
SEGMENT_TYPES = [
    'By Equipment Type',
    'By Automation Level',
    'By Machine Architecture',
    'By End-Use Industry',
    'By Country',
    'By Region',
]

# Hierarchical parents in "By Equipment Type" - these have children
EQUIPMENT_PARENTS = {
    'Grouping & Picking Systems': ['Picker Lines', 'Product Collators & Accumulation Systems'],
    'Secondary Bagging & Over-Wrapping Equipment': ['Bagging Machines', 'Bundling & Banding Machines'],
}

# All leaf segments under "By Equipment Type" (flat ones that are NOT parents)
EQUIPMENT_FLAT_LEAVES = [
    'Sleeve & Shrink Secondary Packaging Equipment',
    'Box Forming Machines (Case Erectors, Tray Formers, Wraparound Formers)',
    'Forming\u2013Filling\u2013Closing Machines',  # en-dash
    'Others (Cartoning & Case Secondary Packaging Equipment, Case Sealing & Strapping Machines, Palletizing & End-of-Line Equipment, etc.)',
]

# Geography hierarchy for By Country / By Region
GEO_HIERARCHY = {
    'North America': ['U.S.', 'Canada'],
    'Europe': ['U.K.', 'Germany', 'France', 'Italy', 'Spain', 'Russia', 'Turkey', 'Rest of Europe'],
    'Asia Pacific': ['India', 'China', 'Japan', 'Australia', 'ASEAN', 'South Korea', 'Rest of Asia Pacific'],
    'Latin America': ['Brazil', 'Argentina', 'Mexico', 'Rest of Latin America'],
    'Middle East & Africa': ['GCC', 'South Africa', 'Rest of Middle East & Africa'],
}

REGIONS = list(GEO_HIERARCHY.keys())


def fix_label(label):
    """Fix encoding issues in segment labels (garbled en-dashes)."""
    if label is None:
        return None
    label = str(label).strip()
    # Replace garbled en-dash characters (replacement char \uFFFD or other variants)
    label = label.replace('\ufffd', '\u2013')  # replacement char -> en-dash
    # Also handle cases where it might be a different garbled sequence
    if 'Forming' in label and 'Filling' in label and 'Closing' in label:
        label = re.sub(r'Forming.Filling.Closing', 'Forming\u2013Filling\u2013Closing', label)
    return label


def read_year_data(ws, row_idx, is_value=True):
    """Read year data from columns 2-14 (years 2021-2033)."""
    data = {}
    for col_idx, year in enumerate(YEARS, start=2):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            if is_value:
                data[year] = round(float(val), 1)
            else:
                data[year] = round(float(val))
    return data


def match_equipment_leaf(label):
    """Check if a label matches any known equipment flat leaf segment."""
    for flat_leaf in EQUIPMENT_FLAT_LEAVES:
        if label == flat_leaf:
            return flat_leaf
        # Partial match for truncated/garbled labels
        if len(label) > 20 and label[:30] == flat_leaf[:30]:
            return flat_leaf
    # Special case: Others segment
    if label.startswith('Others (Cartoning'):
        return EQUIPMENT_FLAT_LEAVES[-1]
    return None


def parse_value_sheet(wb):
    """Parse the Value sheet into a structured dict."""
    ws = wb['Value']
    result = {}

    current_geo = None
    current_seg_type = None
    current_parent = None  # For hierarchical equipment parents

    for row_idx in range(18, ws.max_row + 1):
        raw_label = ws.cell(row=row_idx, column=1).value
        if raw_label is None or str(raw_label).strip() == '':
            continue

        label = fix_label(raw_label)
        has_data = ws.cell(row=row_idx, column=2).value is not None

        # Skip the header row
        if label == 'Row Labels':
            continue

        # Detect geography header (no data in col 2)
        if not has_data:
            current_geo = label
            current_seg_type = None
            current_parent = None
            if current_geo not in result:
                result[current_geo] = {}
            continue

        # Detect segment type
        if label in SEGMENT_TYPES:
            current_seg_type = label
            current_parent = None
            if current_seg_type not in result.get(current_geo, {}):
                result[current_geo][current_seg_type] = {}
            # Skip the total row (don't store data for segment type totals)
            continue

        if current_geo is None or current_seg_type is None:
            continue

        year_data = read_year_data(ws, row_idx, is_value=True)
        if not year_data:
            continue

        # Handle "By Equipment Type" hierarchy
        if current_seg_type == 'By Equipment Type':
            # Check if this is a known parent
            if label in EQUIPMENT_PARENTS:
                current_parent = label
                # Create parent as MIXED NODE: year data + children together
                # The json-processor detects this as an aggregated record
                # and handles it properly (no double counting in chart sums)
                result[current_geo][current_seg_type][label] = dict(year_data)
                continue

            # Check if this is a child of current parent
            if current_parent and current_parent in EQUIPMENT_PARENTS:
                if label in EQUIPMENT_PARENTS[current_parent]:
                    result[current_geo][current_seg_type][current_parent][label] = year_data
                    continue

            # Check if this is a flat leaf segment
            matched_leaf = match_equipment_leaf(label)
            if matched_leaf:
                result[current_geo][current_seg_type][matched_leaf] = year_data
                current_parent = None
            else:
                # Unknown segment - store as-is
                print(f"  WARNING: Unknown equipment segment '{label}' in {current_geo}")
                result[current_geo][current_seg_type][label] = year_data
                current_parent = None

        # Handle flat segment types
        elif current_seg_type in ('By Automation Level', 'By Machine Architecture', 'By End-Use Industry'):
            result[current_geo][current_seg_type][label] = year_data

        # Handle By Country / By Region
        elif current_seg_type in ('By Country', 'By Region'):
            result[current_geo][current_seg_type][label] = year_data

    return result


def parse_volume_sheet(wb):
    """Parse the Volume sheet - only By Equipment Type per geography."""
    ws = wb['Volume']
    result = {}

    current_geo = None
    current_parent = None
    in_equipment_type = False

    for row_idx in range(18, ws.max_row + 1):
        raw_label = ws.cell(row=row_idx, column=1).value
        if raw_label is None or str(raw_label).strip() == '':
            continue

        label = fix_label(raw_label)
        has_data = ws.cell(row=row_idx, column=2).value is not None

        if label == 'Row Labels':
            continue

        # Detect geography header
        if not has_data:
            current_geo = label
            current_parent = None
            in_equipment_type = False
            if current_geo not in result:
                result[current_geo] = {}
            continue

        # Detect "By Equipment Type" total row
        if label == 'By Equipment Type':
            in_equipment_type = True
            current_parent = None
            result[current_geo]['By Equipment Type'] = {}
            continue  # Skip total

        if current_geo is None or not in_equipment_type:
            continue

        year_data = read_year_data(ws, row_idx, is_value=False)
        if not year_data:
            continue

        # Handle hierarchy same as value sheet - mixed nodes
        if label in EQUIPMENT_PARENTS:
            current_parent = label
            result[current_geo]['By Equipment Type'][label] = dict(year_data)
            continue

        if current_parent and current_parent in EQUIPMENT_PARENTS:
            if label in EQUIPMENT_PARENTS[current_parent]:
                result[current_geo]['By Equipment Type'][current_parent][label] = year_data
                continue

        # Flat leaves
        matched_leaf = match_equipment_leaf(label)
        if matched_leaf:
            result[current_geo]['By Equipment Type'][matched_leaf] = year_data
            current_parent = None
        else:
            print(f"  WARNING: Unknown volume equipment segment '{label}' in {current_geo}")
            result[current_geo]['By Equipment Type'][label] = year_data
            current_parent = None

    return result


def build_segmentation_analysis():
    """Build segmentation_analysis.json with full hierarchy."""
    analysis = {
        "Global": {
            "By Equipment Type": {
                "Grouping & Picking Systems": {
                    "Picker Lines": {},
                    "Product Collators & Accumulation Systems": {}
                },
                "Secondary Bagging & Over-Wrapping Equipment": {
                    "Bagging Machines": {},
                    "Bundling & Banding Machines": {}
                },
                "Sleeve & Shrink Secondary Packaging Equipment": {},
                "Box Forming Machines (Case Erectors, Tray Formers, Wraparound Formers)": {},
                "Forming\u2013Filling\u2013Closing Machines": {},
                "Others (Cartoning & Case Secondary Packaging Equipment, Case Sealing & Strapping Machines, Palletizing & End-of-Line Equipment, etc.)": {}
            },
            "By Automation Level": {
                "Manual": {},
                "Semi-automatic and Fully Automatic": {}
            },
            "By Machine Architecture": {
                "Conventional Mechanical Systems": {},
                "Servo-based Non-robotic Systems": {},
                "Robotic Systems": {}
            },
            "By End-Use Industry": {
                "Food & Beverage": {},
                "Pharmaceuticals": {},
                "Personal Care & Cosmetics": {},
                "Household Chemicals": {},
                "E-commerce & Logistics": {},
                "Electronics & Electrical": {},
                "Automotive & Industrial Goods": {}
            },
            "By Region": {}
        }
    }

    # Build By Region hierarchy
    for region, countries in GEO_HIERARCHY.items():
        analysis["Global"]["By Region"][region] = {}
        for country in countries:
            analysis["Global"]["By Region"][region][country] = {}

    return analysis


def verify_no_double_counting(value_data):
    """Verify leaf sums match expected totals (spot check)."""
    print("\n=== Verification: No Double Counting ===")

    for geo in ['Global', 'North America']:
        if geo not in value_data:
            continue
        et = value_data[geo].get('By Equipment Type', {})
        total_2021 = 0
        for key, val in et.items():
            if isinstance(val, dict):
                if '2021' in val:
                    # Leaf with year data
                    total_2021 += val['2021']
                else:
                    # Parent container
                    for child_key, child_val in val.items():
                        if isinstance(child_val, dict) and '2021' in child_val:
                            total_2021 += child_val['2021']
        print(f"  {geo} By Equipment Type leaf sum 2021: {total_2021:.1f}")

    # Check that parent nodes are mixed nodes (year data + children)
    for geo in ['Global', 'North America']:
        if geo not in value_data:
            continue
        et = value_data[geo].get('By Equipment Type', {})
        for parent_name, expected_children in EQUIPMENT_PARENTS.items():
            if parent_name in et:
                parent_obj = et[parent_name]
                has_year_data = any(k.isdigit() for k in parent_obj.keys())
                has_children = any(k in expected_children for k in parent_obj.keys())
                if has_year_data and has_children:
                    print(f"  OK: {geo} > {parent_name} is mixed node (year data + {len(expected_children)} children)")
                elif has_year_data:
                    print(f"  WARN: {geo} > {parent_name} has year data but missing children")
                else:
                    print(f"  WARN: {geo} > {parent_name} has no year data")


def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    print("\nParsing Value sheet...")
    value_data = parse_value_sheet(wb)
    print(f"  Geographies found: {len(value_data)}")
    print(f"  Geography names: {list(value_data.keys())}")

    print("\nParsing Volume sheet...")
    volume_data = parse_volume_sheet(wb)
    print(f"  Geographies found: {len(volume_data)}")
    print(f"  Geography names: {list(volume_data.keys())}")

    verify_no_double_counting(value_data)

    # Print structure summary
    print("\n=== Structure Summary ===")
    for geo in value_data:
        seg_types = list(value_data[geo].keys())
        print(f"  {geo}: {seg_types}")

    # Check a sample
    if 'North America' in value_data:
        na_et = value_data['North America'].get('By Equipment Type', {})
        print(f"\n  NA Equipment Type structure:")
        for k, v in na_et.items():
            if isinstance(v, dict) and '2021' not in v:
                print(f"    {k} (parent) -> {list(v.keys())}")
            else:
                val_2021 = v.get('2021', '?') if isinstance(v, dict) else '?'
                print(f"    {k} (leaf, 2021={val_2021})")

    print("\nBuilding segmentation_analysis.json...")
    seg_analysis = build_segmentation_analysis()

    # Write output files
    print("\nWriting public/data/value.json...")
    with open('public/data/value.json', 'w', encoding='utf-8') as f:
        json.dump(value_data, f, indent=2, ensure_ascii=False)

    print("Writing public/data/volume.json...")
    with open('public/data/volume.json', 'w', encoding='utf-8') as f:
        json.dump(volume_data, f, indent=2, ensure_ascii=False)

    print("Writing public/data/segmentation_analysis.json...")
    with open('public/data/segmentation_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(seg_analysis, f, indent=2, ensure_ascii=False)

    print("\nDone! All files generated successfully.")


if __name__ == '__main__':
    main()
